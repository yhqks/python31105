import openpyxl
from openpyxl.utils import get_column_letter
import pandas as pd

# file=pd.read_excel('test2.xlsx')
# print(file)
import paddle
print(paddle.utils.run_check()) 
# def parse_complex_excel(file_path):
#     # 加载Excel文件
#     workbook = openpyxl.load_workbook(file_path)
#     sheet = workbook.active

#     # 提取合并单元格的范围
#     merged_ranges = []
#     for merged_cell in sheet.merged_cells.ranges:
#         merged_ranges.append(merged_cell)

#     # 解析多级表头（横向和纵向）
#     headers = {}
#     max_header_rows = 0
#     max_header_cols = 0

#     # 检查顶部几行是否是表头
#     for row_idx in range(1, 6):  # 假设表头最多分布在前5行
#         row_headers = []
#         for cell in sheet[row_idx]:
#             # 检查是否是合并单元格
#             is_merged = False
#             merged_value = None
#             for merged_range in merged_ranges:
#                 if cell.coordinate in merged_range:
#                     is_merged = True
#                     # 合并单元格的值取左上角的单元格
#                     merged_start_row = merged_range.min_row
#                     merged_start_col = merged_range.min_col
#                     merged_cell = sheet.cell(row=merged_start_row, column=merged_start_col)
#                     merged_value = merged_cell.value
#                     break
#             if is_merged:
#                 row_headers.append(merged_value)
#             else:
#                 row_headers.append(cell.value)
#         if all(header is None for header in row_headers):
#             break  # 如果某行全是空值，停止解析表头
#         headers[row_idx] = row_headers
#         max_header_rows += 1

#     # 检查左侧几列是否是表头
#     for col_idx in range(1, 6):  # 假设表头最多分布在前5列
#         col_headers = []
#         for row in sheet.iter_rows(min_col=col_idx, max_col=col_idx, min_row=1, max_row=5):
#             cell = row[0]
#             # 检查是否是合并单元格
#             is_merged = False
#             merged_value = None
#             for merged_range in merged_ranges:
#                 if cell.coordinate in merged_range:
#                     is_merged = True
#                     # 合并单元格的值取左上角的单元格
#                     merged_start_row = merged_range.min_row
#                     merged_start_col = merged_range.min_col
#                     merged_cell = sheet.cell(row=merged_start_row, column=merged_start_col)
#                     merged_value = merged_cell.value
#                     break
#             if is_merged:
#                 col_headers.append(merged_value)
#             else:
#                 col_headers.append(cell.value)
#         if all(header is None for header in col_headers):
#             break  # 如果某列全是空值，停止解析表头
#         headers[col_idx] = col_headers
#         max_header_cols += 1

#     # 解析数据行
#     data = []
#     for row_idx, row in enumerate(sheet.iter_rows(min_row=max_header_rows + 1), start=1):
#         data_row = {}
#         for col_idx, cell in enumerate(row, start=1):
#             # 检查是否是合并单元格
#             is_merged = False
#             merged_value = None
#             for merged_range in merged_ranges:
#                 if cell.coordinate in merged_range:
#                     is_merged = True
#                     # 合并单元格的值取左上角的单元格
#                     merged_start_row = merged_range.min_row
#                     merged_start_col = merged_range.min_col
#                     merged_cell = sheet.cell(row=merged_start_row, column=merged_start_col)
#                     merged_value = merged_cell.value
#                     break

#             # 获取表头信息
#             header = None
#             # 优先检查横向表头
#             if max_header_rows > 0 and col_idx - 1 < len(headers[max_header_rows]):
#                 header = headers[max_header_rows][col_idx - 1]
#             # 检查纵向表头
#             if max_header_cols > 0 and header is None and row_idx - max_header_rows - 1 < len(headers[col_idx]):
#                 header = headers[col_idx][row_idx - max_header_rows - 1]

#             if header is None:
#                 header = f"Column_{col_idx}"  # 如果没有表头，使用默认名称

#             if is_merged:
#                 data_row[header] = merged_value
#             else:
#                 data_row[header] = cell.value
#         data.append(data_row)

#     return headers, data

# # 示例：解析Excel文件
# file_path = 'test2.xlsx'
# headers, data = parse_complex_excel(file_path)

# # 打印表头
# print("表头:")
# for key, value in headers.items():
#     print(f"Row/Col {key}: {value}")

# # 打印数据
# print("\n数据:")
# for row in data:
#     print(row)